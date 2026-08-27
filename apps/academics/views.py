from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from .models import (
    AcademicProgram, Exam, ExamRoutine, Result, SubjectMark,
    TeacherSubjectAssignment, MarksAuditLog,
)
from apps.core.models import SchoolSettings
from .forms import ResultSearchForm, ExamRoutineBulkUploadForm, ResultEditForm
from .services import save_marks_and_result, calculate_result_from_subject_marks
from .permissions import (
    can_enter_marks, can_edit_result, validate_exam_exists,
    validate_grade, validate_subject_mark_data,
    is_teacher_assigned_to_subject, get_authorized_subjects,
)
from django.db import transaction
from datetime import datetime

import io
import csv
from django.contrib.admin.views.decorators import staff_member_required


def programs(request):
    """List all academic programs."""
    school = SchoolSettings.get_settings()

    # ── Redis cache (6 hours) ──
    cache_key = 'galaxy:programs:list'
    program_list = cache.get(cache_key)
    if program_list is None:
        program_list = list(
            AcademicProgram.objects.filter(is_published=True)
            .order_by('display_order')
        )
        cache.set(cache_key, program_list, 6 * 60 * 60)

    context = {
        'school': school,
        'programs': program_list,
        'page_title': f"Academic Programs - {school.school_name}",
        'meta_description': f"Explore academic programs offered at {school.school_name}, Bhadrapur-9, Jhapa, Nepal.",
    }
    return render(request, 'academics/programs.html', context)


def program_detail(request, slug):
    """Academic program detail page."""
    school = SchoolSettings.get_settings()
    program = get_object_or_404(
        AcademicProgram,
        slug=slug,
        is_published=True
    )
    other_programs = AcademicProgram.objects.filter(
        is_published=True
    ).exclude(pk=program.pk).order_by('display_order')[:4]

    context = {
        'school': school,
        'program': program,
        'other_programs': other_programs,
        'page_title': f"{program.name} - {school.school_name}",
        'meta_description': program.short_description,
    }
    return render(request, 'academics/program_detail.html', context)


def exam_routine(request, exam_id=None):
    """Exam routine page.

    * Only published examinations are ever shown publicly.
    * Users can select an examination and a class (Class 1 - Class 10).
    * Routines are filtered with the ORM and sorted by exam date then
      start time. Unpublished exams are never exposed.
    * Results cached in Redis for 30 minutes.
    """
    school = SchoolSettings.get_settings()

    selected_exam_id = exam_id or request.GET.get('exam')
    selected_grade = request.GET.get('grade', '')

    # Only published examinations are visible publicly.
    published_exams = Exam.objects.filter(is_published=True).order_by(
        '-start_date'
    ).only('id', 'name', 'academic_year', 'start_date')

    selected_exam = None
    if selected_exam_id:
        selected_exam = published_exams.filter(pk=selected_exam_id).first()
    if selected_exam is None:
        selected_exam = published_exams.first()

    # Prep the queryset using ORM filtering (never manual template filtering).
    routines = ExamRoutine.objects.none()
    # Group routines into one table per class so classes are never merged.
    class_routines = []
    if selected_exam:
        # ── Redis cache (30 min) ──
        cache_key = f"galaxy:routine:{selected_exam.pk}:{selected_grade or 'all'}"
        class_routines = cache.get(cache_key)

        if class_routines is None:
            routine_qs = ExamRoutine.objects.filter(
                exam=selected_exam
            ).order_by('grade', 'exam_date', 'start_time')

            # Filter by the selected class (if any).
            if selected_grade:
                routine_qs = routine_qs.filter(grade=selected_grade)

            routines = routine_qs

            # Build one group per class, keeping Class 1 -> Class 10 order and
            # preserving the date/time ordering inside each class.
            grade_label_map = dict(ExamRoutine.GRADE_CHOICES)
            grade_order = [value for value, _ in ExamRoutine.GRADE_CHOICES]
            graded = {value: [] for value in grade_order}
            for routine in routines:
                graded.setdefault(routine.grade, []).append(routine)

            class_routines = [
                {
                    'grade': value,
                    'grade_label': grade_label_map[value],
                    'routines': graded[value],
                }
                for value in grade_order
                if graded.get(value)
            ]
            cache.set(cache_key, class_routines, 30 * 60)  # 30 min TTL
        # On cache hit, routines queryset is not needed —
        # the template only uses class_routines for display.

    context = {
        'school': school,
        'published_exams': published_exams,
        'selected_exam': selected_exam,
        'routines': routines,
        'class_routines': class_routines,
        'selected_grade': selected_grade,
        'grade_choices': ExamRoutine.GRADE_CHOICES,
        'page_title': f"Exam Routine - {school.school_name}",
        'meta_description': f"View exam routines for Classes 1-10 at {school.school_name}.",
    }
    return render(request, 'academics/exam_routine.html', context)
from apps.core.ratelimit import rate_limit


@rate_limit(
    key_prefix='result_search',
    max_requests=30,
    window_seconds=60,  # 30 requests per minute
    block_message='Too many search requests. Please wait a moment before trying again.',
)
def results(request):
    """Result search page.

    Results are cached in Redis for 5 minutes per symbol_number + exam.
    Cache is invalidated automatically when a result is saved/edited
    (see tasks.py and the save_marks_and_result service).
    """
    school = SchoolSettings.get_settings()
    form = ResultSearchForm()
    result_obj = None
    searched = False
    error_message = ""

    if request.method == 'GET' and request.GET.get('symbol_number'):
        form = ResultSearchForm(request.GET)
        searched = True

        if form.is_valid():
            symbol_number = form.cleaned_data['symbol_number']
            exam_id = form.cleaned_data.get('exam')

            # ── Redis cache (5 min) ──
            cache_key = f"galaxy:result:{symbol_number}:{exam_id or 'latest'}"
            cached_result = cache.get(cache_key)

            if cached_result is not None:
                # Cache hit — reconstruct a lightweight object from the dict
                # We still need the full ORM object for the template, so cache
                # the result_id and query from there.
                result_obj = cached_result  # stored as ORM pk
                if isinstance(cached_result, int):
                    try:
                        result_obj = Result.objects.select_related('exam').prefetch_related('subject_marks').get(pk=cached_result)
                    except Result.DoesNotExist:
                        result_obj = None
                        cache.delete(cache_key)
            else:
                try:
                    query = Result.objects.filter(
                        symbol_number__iexact=symbol_number,
                        is_published=True
                    ).select_related('exam').prefetch_related('subject_marks')

                    if exam_id:
                        query = query.filter(exam_id=exam_id)

                    result_obj = query.latest('published_at')
                    # Cache the result ID for 5 minutes
                    cache.set(cache_key, result_obj.pk, 5 * 60)

                except Result.DoesNotExist:
                    error_message = f"No published result found for symbol number '{symbol_number}'. Please verify your symbol number and try again."

    # Published exams with results
    available_exams = Exam.objects.filter(
        is_published=True,
        results__is_published=True
    ).distinct().order_by('-start_date')

    context = {
        'school': school,
        'form': form,
        'result': result_obj,
        'searched': searched,
        'error_message': error_message,
        'available_exams': available_exams,
        'page_title': f"Check Results - {school.school_name}",
        'meta_description': f"Check your exam results at {school.school_name}, Bhadrapur-9, Jhapa, Nepal.",
    }
    return render(request, 'academics/results.html', context)
def parse_date(value):
    if not value:
        raise ValueError("Exam date is required.")

    # Excel datetime/date
    if hasattr(value, "date"):
        return value.date()

    value = str(value).strip()

    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
    ]

    for date_format in formats:
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            pass

    raise ValueError(
        f"Invalid date '{value}'. Use YYYY-MM-DD."
    )


def parse_time(value):
    if not value:
        raise ValueError("Time is required.")

    # Excel time/datetime
    if hasattr(value, "time"):
        return value.time()

    value = str(value).strip()

    formats = [
        "%H:%M",
        "%H:%M:%S",
        "%I:%M %p",
        "%I:%M:%S %p",
    ]

    for time_format in formats:
        try:
            return datetime.strptime(
                value,
                time_format
            ).time()
        except ValueError:
            pass

    raise ValueError(
        f"Invalid time '{value}'. Use HH:MM."
    )


def read_csv_file(uploaded_file):

    content = uploaded_file.read().decode("utf-8-sig")

    reader = csv.DictReader(
        io.StringIO(content)
    )

    return list(reader)


def read_excel_file(uploaded_file):
    from openpyxl import load_workbook

    workbook = load_workbook(
        uploaded_file,
        read_only=True,
        data_only=True
    )

    worksheet = workbook.active

    rows = worksheet.iter_rows(
        values_only=True
    )

    try:
        headers = next(rows)
    except StopIteration:
        return []

    headers = [
        str(header).strip().lower()
        if header is not None
        else ""
        for header in headers
    ]

    result = []

    for row in rows:

        if not any(value is not None for value in row):
            continue

        row_data = {}

        for index, header in enumerate(headers):

            if not header:
                continue

            row_data[header] = (
                row[index]
                if index < len(row)
                else None
            )

        result.append(row_data)

    return result


def normalize_row(row):

    return {
        str(key)
        .strip()
        .lower()
        .replace(" ", "_"): value

        for key, value in row.items()
        if key is not None
    }


@staff_member_required
def bulk_upload_exam_routine(request, exam_id):

    exam = get_object_or_404(
        Exam,
        pk=exam_id
    )

    if request.method == "POST":

        form = ExamRoutineBulkUploadForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            uploaded_file = form.cleaned_data["file"]

            try:

                filename = uploaded_file.name.lower()

                if filename.endswith(".csv"):
                    rows = read_csv_file(
                        uploaded_file
                    )

                else:
                    rows = read_excel_file(
                        uploaded_file
                    )

                if not rows:
                    raise ValueError(
                        "The uploaded file contains no data."
                    )

                required_columns = {
                    "grade",
                    "subject",
                    "exam_date",
                    "start_time",
                    "end_time",
                }

                first_row = normalize_row(rows[0])

                missing_columns = (
                    required_columns
                    - set(first_row.keys())
                )

                if missing_columns:

                    raise ValueError(
                        "Missing required columns: "
                        + ", ".join(
                            sorted(missing_columns)
                        )
                    )

                routines = []
                errors = []

                valid_grades = dict(
                    ExamRoutine.GRADE_CHOICES
                )

                for row_number, raw_row in enumerate(
                    rows,
                    start=2
                ):

                    row = normalize_row(raw_row)

                    try:

                        grade = str(
                            row.get("grade", "")
                        ).strip()

                        subject = str(
                            row.get("subject", "")
                        ).strip()

                        room = str(
                            row.get("room", "")
                            or ""
                        ).strip()

                        remarks = str(
                            row.get("remarks", "")
                            or ""
                        ).strip()

                        if grade not in valid_grades:
                            raise ValueError(
                                f"Invalid grade '{grade}'. "
                                "Use 1 to 10."
                            )

                        if not subject:
                            raise ValueError(
                                "Subject is required."
                            )

                        exam_date = parse_date(
                            row.get("exam_date")
                        )

                        start_time = parse_time(
                            row.get("start_time")
                        )

                        end_time = parse_time(
                            row.get("end_time")
                        )

                        if start_time >= end_time:
                            raise ValueError(
                                "Start time must be before "
                                "end time."
                            )

                        routines.append(
                            ExamRoutine(
                                exam=exam,
                                grade=grade,
                                subject=subject,
                                exam_date=exam_date,
                                start_time=start_time,
                                end_time=end_time,
                                room=room,
                                remarks=remarks,
                            )
                        )

                    except Exception as error:

                        errors.append(
                            f"Row {row_number}: {error}"
                        )

                # Don't import anything if any row is invalid
                if errors:

                    return render(
                        request,
                        "academics/exam/"
                        "bulk_upload_routine.html",
                        {
                            "form": form,
                            "exam": exam,
                            "errors": errors,
                        }
                    )

                # Prevent duplicates
                for routine in routines:

                    exists = ExamRoutine.objects.filter(
                        exam=exam,
                        grade=routine.grade,
                        subject=routine.subject,
                        exam_date=routine.exam_date,
                        start_time=routine.start_time,
                    ).exists()

                    if exists:

                        errors.append(
                            f"Duplicate found: "
                            f"Class {routine.grade} - "
                            f"{routine.subject} - "
                            f"{routine.exam_date}"
                        )

                if errors:

                    return render(
                        request,
                        "academics/exam/"
                        "bulk_upload_routine.html",
                        {
                            "form": form,
                            "exam": exam,
                            "errors": errors,
                        }
                    )

                # Atomic import
                with transaction.atomic():

                    ExamRoutine.objects.bulk_create(
                        routines,
                        batch_size=500
                    )

                # ── Invalidate routine caches ──
                cache.delete(f"galaxy:routine:{exam.pk}:all")
                try:
                    from .tasks import invalidate_routine_caches
                    invalidate_routine_caches.delay(exam.pk)
                except Exception:
                    pass

                messages.success(
                    request,
                    f"Successfully imported "
                    f"{len(routines)} exam routines."
                )

                return redirect(
                    "admin:academics_exam_change",
                    object_id=exam.pk
                )

            except Exception as error:

                form.add_error(
                    "file",
                    str(error)
                )

    else:

        form = ExamRoutineBulkUploadForm()

    return render(
        request,
        "academics/exam/"
        "bulk_upload_routine.html",
        {
            "form": form,
            "exam": exam,
        }
    )


@staff_member_required
def exam_routine_add(request, exam_id=None):
    """Admin page to add exam routines by class.

    Only accessible by superusers (admins). Teachers can only add marks/results.
    """
    if not request.user.is_superuser:
        messages.error(request, "You don't have permission to add exam routines. Only administrators can access this page.")
        return redirect('core:home')
    school = SchoolSettings.get_settings()

    exams = Exam.objects.all().order_by('-start_date')
    selected_exam = None
    selected_grade = request.GET.get('grade', '1')
    routines = []
    manual_form_errors = []

    if exam_id:
        selected_exam = get_object_or_404(Exam, pk=exam_id)
    elif request.GET.get('exam'):
        selected_exam = get_object_or_404(
            Exam, pk=request.GET.get('exam')
        )
    elif exams.exists():
        selected_exam = exams.first()

    # ── Manual single-routine entry (easy form) ──
    if request.method == 'POST':
        if request.POST.get('action') == 'add_manual':
            if not selected_exam:
                messages.error(request, "Please select an exam first.")
            else:
                subject = (request.POST.get('subject') or '').strip()
                exam_date_raw = (request.POST.get('exam_date') or '').strip()
                start_time_raw = (request.POST.get('start_time') or '').strip()
                end_time_raw = (request.POST.get('end_time') or '').strip()
                room = (request.POST.get('room') or '').strip()
                remarks = (request.POST.get('remarks') or '').strip()
                manual_grade = (request.POST.get('grade') or selected_grade or '').strip()

                try:
                    if not subject:
                        raise ValueError("Subject is required.")
                    if not exam_date_raw:
                        raise ValueError("Exam date is required.")
                    if not start_time_raw or not end_time_raw:
                        raise ValueError("Start and end time are required.")

                    try:
                        exam_date = datetime.strptime(
                            exam_date_raw, "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        raise ValueError("Invalid date format.")

                    try:
                        start_time = datetime.strptime(
                            start_time_raw, "%H:%M"
                        ).time()
                        end_time = datetime.strptime(
                            end_time_raw, "%H:%M"
                        ).time()
                    except ValueError:
                        raise ValueError(
                            "Invalid time format. Use HH:MM (24h)."
                        )

                    if start_time >= end_time:
                        raise ValueError(
                            "Start time must be before end time."
                        )

                    valid_grades = dict(ExamRoutine.GRADE_CHOICES)
                    if manual_grade not in valid_grades:
                        raise ValueError("Please choose a valid class.")

                    if ExamRoutine.objects.filter(
                        exam=selected_exam,
                        grade=manual_grade,
                        subject=subject,
                        exam_date=exam_date,
                        start_time=start_time,
                    ).exists():
                        raise ValueError(
                            "A routine with the same class, subject, "
                            "date and start time already exists."
                        )

                    ExamRoutine.objects.create(
                        exam=selected_exam,
                        grade=manual_grade,
                        subject=subject,
                        exam_date=exam_date,
                        start_time=start_time,
                        end_time=end_time,
                        room=room,
                        remarks=remarks,
                    )
                    # ── Invalidate routine cache ──
                    cache.delete(f"galaxy:routine:{selected_exam.pk}:all")
                    cache.delete(f"galaxy:routine:{selected_exam.pk}:{manual_grade}")
                    messages.success(
                        request,
                        f"Routine added: Class {manual_grade} - "
                        f"{subject} on {exam_date}."
                    )
                    # Refresh selected_grade so the new row is visible
                    selected_grade = manual_grade

                except ValueError as error:
                    manual_form_errors.append(str(error))
                    messages.error(request, str(error))

        # ── Class-wise routine entry (add all subjects for one class) ──
        elif request.POST.get('action') == 'add_class_routine':
            if not selected_exam:
                messages.error(request, "Please select an exam first.")
            else:
                # Get form data arrays
                subjects = request.POST.getlist('subject[]')
                start_times = request.POST.getlist('start_time[]')
                end_times = request.POST.getlist('end_time[]')
                rooms = request.POST.getlist('room[]')
                remarks = request.POST.getlist('remarks[]')
                manual_grade = (request.POST.get('grade') or selected_grade or '').strip()
                exam_date_raw = (request.POST.get('exam_date') or '').strip()

                try:
                    if not manual_grade:
                        raise ValueError("Please select a class.")
                    if not exam_date_raw:
                        raise ValueError("Exam date is required.")
                    
                    try:
                        exam_date = datetime.strptime(
                            exam_date_raw, "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        raise ValueError("Invalid date format. Use YYYY-MM-DD.")
                    
                    valid_grades = dict(ExamRoutine.GRADE_CHOICES)
                    if manual_grade not in valid_grades:
                        raise ValueError("Please choose a valid class.")
                    
                    # Check if we have at least one subject
                    if not subjects or all(not s.strip() for s in subjects):
                        raise ValueError("Please enter at least one subject.")
                    
                    # Process each subject
                    created_count = 0
                    errors = []
                    
                    for i, subject in enumerate(subjects):
                        subject = subject.strip()
                        if not subject:
                            continue  # Skip empty subjects
                        
                        start_time_str = start_times[i] if i < len(start_times) else ''
                        end_time_str = end_times[i] if i < len(end_times) else ''
                        room_str = rooms[i] if i < len(rooms) else ''
                        remark_str = remarks[i] if i < len(remarks) else ''
                        
                        if not start_time_str or not end_time_str:
                            errors.append(f"Row {i+1}: Start and end time are required for '{subject}'.")
                            continue
                        
                        try:
                            start_time = datetime.strptime(start_time_str, "%H:%M").time()
                            end_time = datetime.strptime(end_time_str, "%H:%M").time()
                        except ValueError:
                            errors.append(f"Row {i+1}: Invalid time format for '{subject}'. Use HH:MM.")
                            continue
                        
                        if start_time >= end_time:
                            errors.append(f"Row {i+1}: Start time must be before end time for '{subject}'.")
                            continue
                        
                        # Check for duplicate
                        if ExamRoutine.objects.filter(
                            exam=selected_exam,
                            grade=manual_grade,
                            subject=subject,
                            exam_date=exam_date,
                            start_time=start_time,
                        ).exists():
                            errors.append(f"Row {i+1}: Duplicate routine for '{subject}' at {start_time}.")
                            continue
                        
                        # Create the routine
                        ExamRoutine.objects.create(
                            exam=selected_exam,
                            grade=manual_grade,
                            subject=subject,
                            exam_date=exam_date,
                            start_time=start_time,
                            end_time=end_time,
                            room=room_str.strip(),
                            remarks=remark_str.strip(),
                        )
                        created_count += 1
                    
                    # ── Invalidate routine cache ──
                    cache.delete(f"galaxy:routine:{selected_exam.pk}:all")
                    cache.delete(f"galaxy:routine:{selected_exam.pk}:{manual_grade}")

                    if errors:
                        for error in errors:
                            manual_form_errors.append(error)
                            messages.error(request, error)
                    else:
                        messages.success(
                            request,
                            f"Successfully added {created_count} routines for Class {manual_grade} on {exam_date}."
                        )
                        # Refresh selected_grade so the new rows are visible
                        selected_grade = manual_grade
                        
                except ValueError as error:
                    manual_form_errors.append(str(error))
                    messages.error(request, str(error))

    # Refresh routines list for the current exam + grade
    if selected_exam:
        routines = ExamRoutine.objects.filter(
            exam=selected_exam
        ).order_by('grade', 'exam_date', 'start_time')
        if selected_grade:
            routines = routines.filter(grade=selected_grade)

    context = {
        'school': school,
        'exams': exams,
        'selected_exam': selected_exam,
        'selected_grade': selected_grade,
        'routines': routines,
        'grade_choices': ExamRoutine.GRADE_CHOICES,
        'manual_form_errors': manual_form_errors,
        'page_title': 'Add Exam Routine',
    }
    return render(
        request,
        'academics/admin_exam_routine.html',
        context
    )


@staff_member_required
def exam_result_add(request):
    """Admin page to add or edit exam results.

    Security:
    - Only staff/superusers can access.
    - GPA/percentage/total/status are NEVER trusted from the client.
    - All calculations happen server-side via services.py.
    - Teacher subject assignments are checked.
    """
    school = SchoolSettings.get_settings()

    exams = Exam.objects.all().order_by('-start_date')
    selected_exam = None
    selected_grade = request.GET.get('grade', '1')
    results = []
    manual_form_errors = []
    editing_result = None
    editing_subject_marks = []

    # ── Load selected exam ──
    if request.GET.get('exam'):
        selected_exam = get_object_or_404(
            Exam, pk=request.GET.get('exam')
        )

    # ── Load editing result if requested ──
    edit_id = request.GET.get('edit')
    if edit_id:
        try:
            editing_result = Result.objects.prefetch_related('subject_marks').get(pk=edit_id)
            if not can_edit_result(request.user, editing_result):
                messages.error(request, "You don't have permission to edit this result.")
                editing_result = None
            else:
                editing_subject_marks = list(editing_result.subject_marks.all())
                selected_exam = editing_result.exam
                selected_grade = editing_result.grade
        except Result.DoesNotExist:
            messages.error(request, "Result not found.")

    # ── Handle POST: Add new result ──
    if request.method == 'POST' and request.POST.get('action') == 'add_manual':
        if not can_enter_marks(request.user):
            messages.error(request, "You don't have permission to enter marks.")
        elif not selected_exam:
            messages.error(request, "Please select an exam first.")
        else:
            student_name = (request.POST.get('student_name') or '').strip()
            symbol_number = (request.POST.get('symbol_number') or '').strip()
            manual_grade = (request.POST.get('grade') or selected_grade or '').strip()
            roll_number = (request.POST.get('roll_number') or '').strip()

            # IGNORE client-submitted: gpa, percentage, total_marks, result_status
            # These will be calculated server-side

            # Validate grade
            grade_valid, grade_error = validate_grade(manual_grade)
            if not grade_valid:
                manual_form_errors.append(grade_error)
                messages.error(request, grade_error)
            elif not student_name:
                manual_form_errors.append("Student name is required.")
                messages.error(request, "Student name is required.")
            elif not symbol_number:
                manual_form_errors.append("Symbol number is required.")
                messages.error(request, "Symbol number is required.")
            elif Result.objects.filter(
                exam=selected_exam,
                symbol_number__iexact=symbol_number
            ).exists():
                msg = f"A result for symbol number '{symbol_number}' already exists for this exam."
                manual_form_errors.append(msg)
                messages.error(request, msg)
            else:
                # Parse subject marks from POST
                subjects = request.POST.getlist('subject[]')
                full_marks_list = request.POST.getlist('full_marks[]')
                obtained_marks_list = request.POST.getlist('obtained_marks[]')

                # Validate subject marks data
                subject_data, validation_errors = validate_subject_mark_data(
                    subjects, full_marks_list, obtained_marks_list
                )

                if validation_errors:
                    for err in validation_errors:
                        manual_form_errors.append(err)
                        messages.error(request, err)
                else:
                    # ── Server-side teacher-subject authorization ──
                    # Superusers can enter any subject. Staff teachers can only
                    # enter marks for subjects they are assigned to.
                    auth_errors = []
                    if not request.user.is_superuser:
                        for sd in subject_data:
                            subj_name = sd.get("subject", "").strip()
                            if subj_name and not is_teacher_assigned_to_subject(
                                request.user, manual_grade, subj_name
                            ):
                                auth_errors.append(
                                    f"You are not authorized to enter marks for '{subj_name}' "
                                    f"in Class {manual_grade}."
                                )
                    if auth_errors:
                        for err in auth_errors:
                            manual_form_errors.append(err)
                            messages.error(request, err)
                    else:
                        try:
                            # Create result first (with placeholder values — service will update)
                            result = Result.objects.create(
                                exam=selected_exam,
                                student_name=student_name,
                                symbol_number=symbol_number,
                                grade=manual_grade,
                                roll_number=roll_number,
                                result_status='PASS',  # Will be recalculated
                                is_published=True,
                                published_at=timezone.now(),
                            )

                            # Use service layer to save marks and recalculate
                            result = save_marks_and_result(
                                result=result,
                                subject_marks_data=subject_data,
                                user=request.user,
                                notes="Initial entry via admin form.",
                            )

                            # ── Invalidate result cache ──
                            cache.delete(f"galaxy:result:{symbol_number}:latest")
                            cache.delete(f"galaxy:result:{symbol_number}:{selected_exam.pk}")
                            try:
                                from .tasks import invalidate_result_cache
                                invalidate_result_cache.delay(symbol_number, selected_exam.pk)
                            except Exception:
                                pass  # Celery may not be running — cache.delete above suffices

                            messages.success(
                                request,
                                f"Result added: {student_name} ({symbol_number}). "
                                f"GPA: {result.gpa}, Status: {result.get_result_status_display()}"
                            )
                            selected_grade = manual_grade

                        except Exception as e:
                            manual_form_errors.append(str(e))
                            messages.error(request, str(e))

    # ── Handle POST: Edit existing result ──
    elif request.method == 'POST' and request.POST.get('action') == 'edit_result':
        result_id = request.POST.get('result_id')
        if not result_id:
            messages.error(request, "No result specified for editing.")
        else:
            try:
                result = Result.objects.get(pk=result_id)
            except Result.DoesNotExist:
                messages.error(request, "Result not found.")
            else:
                if not can_edit_result(request.user, result):
                    messages.error(request, "You don't have permission to edit this result.")
                else:
                    student_name = (request.POST.get('student_name') or '').strip()
                    symbol_number = (request.POST.get('symbol_number') or '').strip()
                    manual_grade = (request.POST.get('grade') or result.grade).strip()
                    roll_number = (request.POST.get('roll_number') or '').strip()

                    # Validate
                    grade_valid, grade_error = validate_grade(manual_grade)
                    if not grade_valid:
                        manual_form_errors.append(grade_error)
                        messages.error(request, grade_error)
                    elif not student_name:
                        manual_form_errors.append("Student name is required.")
                    elif not symbol_number:
                        manual_form_errors.append("Symbol number is required.")
                    else:
                        # Check duplicate symbol (excluding current result)
                        if Result.objects.filter(
                            exam=result.exam,
                            symbol_number__iexact=symbol_number,
                        ).exclude(pk=result.pk).exists():
                            msg = f"A result for symbol number '{symbol_number}' already exists for this exam."
                            manual_form_errors.append(msg)
                            messages.error(request, msg)
                        else:
                            # Parse subject marks
                            subjects = request.POST.getlist('subject[]')
                            full_marks_list = request.POST.getlist('full_marks[]')
                            obtained_marks_list = request.POST.getlist('obtained_marks[]')

                            subject_data, validation_errors = validate_subject_mark_data(
                                subjects, full_marks_list, obtained_marks_list
                            )

                            if validation_errors:
                                for err in validation_errors:
                                    manual_form_errors.append(err)
                                    messages.error(request, err)
                            else:
                                # ── Server-side teacher-subject authorization ──
                                auth_errors = []
                                if not request.user.is_superuser:
                                    for sd in subject_data:
                                        subj_name = sd.get("subject", "").strip()
                                        target_grade = manual_grade or result.grade
                                        if subj_name and not is_teacher_assigned_to_subject(
                                            request.user, target_grade, subj_name
                                        ):
                                            auth_errors.append(
                                                f"You are not authorized to enter marks for '{subj_name}' "
                                                f"in Class {target_grade}."
                                            )
                                if auth_errors:
                                    for err in auth_errors:
                                        manual_form_errors.append(err)
                                        messages.error(request, err)
                                else:
                                    try:
                                        # Update student info
                                        result.student_name = student_name
                                        result.symbol_number = symbol_number
                                        result.grade = manual_grade
                                        result.roll_number = roll_number
                                        result.save()

                                        # Use service layer to update marks and recalculate
                                        result = save_marks_and_result(
                                            result=result,
                                            subject_marks_data=subject_data,
                                            user=request.user,
                                            notes="Edited via admin form.",
                                        )

                                        # ── Invalidate result cache ──
                                        cache.delete(f"galaxy:result:{symbol_number}:latest")
                                        cache.delete(f"galaxy:result:{symbol_number}:{result.exam.pk}")
                                        try:
                                            from .tasks import invalidate_result_cache
                                            invalidate_result_cache.delay(symbol_number, result.exam.pk)
                                        except Exception:
                                            pass

                                        messages.success(
                                            request,
                                            f"Result updated: {student_name} ({symbol_number}). "
                                            f"GPA: {result.gpa}, Status: {result.get_result_status_display()}"
                                        )
                                        selected_grade = manual_grade
                                        editing_result = None  # Exit edit mode

                                    except Exception as e:
                                        manual_form_errors.append(str(e))
                                        messages.error(request, str(e))

    # ── Refresh results list ──
    if selected_exam:
        results = Result.objects.filter(
            exam=selected_exam
        ).select_related('exam').prefetch_related('subject_marks').order_by('-published_at')
        if selected_grade:
            results = results.filter(grade=selected_grade)

    # ── Compute authorized subjects for the teacher ──
    authorized_subjects = get_authorized_subjects(request.user, selected_grade) if selected_grade else []

    context = {
        'school': school,
        'exams': exams,
        'selected_exam': selected_exam,
        'selected_grade': selected_grade,
        'results': results,
        'grade_choices': ExamRoutine.GRADE_CHOICES,
        'result_status_choices': Result.RESULT_STATUS_CHOICES,
        'manual_form_errors': manual_form_errors,
        'editing_result': editing_result,
        'editing_subject_marks': editing_subject_marks,
        'authorized_subjects': authorized_subjects,
        'page_title': 'Add Exam Result',
    }
    return render(
        request,
        'academics/admin_exam_result.html',
        context
    )


@login_required
def exam_result_audit_log(request, result_id):
    """View audit log for a specific result. Staff only."""
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to view audit logs.")
        return redirect('academics:exam_result_add')

    result = get_object_or_404(Result, pk=result_id)
    audit_logs = MarksAuditLog.objects.filter(
        result=result
    ).select_related('changed_by').order_by('-created_at')

    context = {
        'result': result,
        'audit_logs': audit_logs,
        'page_title': f'Audit Log - {result.student_name}',
    }
    return render(
        request,
        'academics/exam_result_audit_log.html',
        context
    )